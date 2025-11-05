from odoo import models, fields, api
from odoo.exceptions import UserError
from openpyxl import load_workbook
import base64
import io
import datetime


class HrEmployeeEvaluation(models.Model):
    _name = "hr.employee.evaluation"
    _description = "Employee Evaluation"
    _order = "date desc, id desc"

    employee_id = fields.Many2one("hr.employee", required=True, ondelete="cascade")
    date = fields.Date(default=fields.Date.context_today, required=True)
    line_ids = fields.One2many("hr.employee.evaluation.line", "evaluation_id", copy=True)

    total = fields.Integer(compute="_compute_score", store=True)
    average = fields.Float(compute="_compute_score", digits=(16, 2), store=True)

    note = fields.Text()
    final_score = fields.Selection(
        [
            ("foarte_slaba", "Foarte slabă (20–50 puncte)"),
            ("slaba", "Slabă (50–70 puncte)"),
            ("medie", "Medie (60–70 puncte)"),
            ("buna", "Bună (70–85 puncte)"),
            ("foarte_buna", "Foarte bună (peste 85 puncte)"),
        ],
        compute="_compute_final_score",
        store=True,
    )

    export_file = fields.Binary("Exported File", readonly=True)
    export_filename = fields.Char("Export Filename")

    @api.depends("line_ids.score")
    def _compute_score(self):
        for rec in self:
            nums = [l.score for l in rec.line_ids if l.score]
            rec.total = sum(nums)
            rec.average = (sum(nums) / len(nums)) if nums else 0.0

    @api.depends("total")
    def _compute_final_score(self):
        for rec in self:
            if rec.total <= 50:
                rec.final_score = "foarte_slaba"
            elif rec.total <= 70:
                rec.final_score = "slaba"
            elif rec.total <= 80:
                rec.final_score = "medie"
            elif rec.total <= 85:
                rec.final_score = "buna"
            else:
                rec.final_score = "foarte_buna"


    @api.model
    def create(self, vals):
        line_ids = vals.get("line_ids", [])
        for cmd in line_ids:
            if cmd[0] == 0:
                line_vals = cmd[2]
                if not line_vals.get("question_id"):
                    answer_id = line_vals.get("answer_id")
                    if answer_id:
                        answer = self.env['hr.employee.evaluation.answer'].browse(answer_id)
                        if answer.exists():
                            line_vals["question_id"] = answer.question_id.id
                if not line_vals.get("question_id"):
                    raise UserError("Each evaluation line must have a question.")
        return super().create(vals)



    @api.model
    def default_get(self, fields_list):
        """Auto-create one line per question and set employee automatically."""
        res = super().default_get(fields_list)

        if self.env.context.get("active_model") == "hr.employee" and self.env.context.get("active_id"):
            employee = self.env['hr.employee'].browse(self.env.context['active_id'])
        else:
            employee = self.env['hr.employee'].search([('user_id', '=', self.env.uid)], limit=1)

        if employee:
            res['employee_id'] = employee.id

            if employee.user_id.angajat_function == 'executie':
                questions = self.env["hr.employee.evaluation.question"].search(
                    [("active", "=", True)], order="sequence"
                )
            else:
                questions = self.env["hr.employee.evaluation.question.lead"].search(
                    [("active", "=", True)], order="sequence"
                )
            res["line_ids"] = [(0, 0, {"question_id": q.id}) for q in questions]

        return res

    def action_export_excel(self):
        """Export evaluation results to Excel template."""
        self.ensure_one()

        template_path = self.env["ir.config_parameter"].sudo().get_param(
            "hr_employee_evaluation.template_path",
            default="/odoo-18/custom-addons/wisdom_employee_review/static/template/Chestionar-evaluare-salariati-executie.xlsx",
        )
        try:
            wb = load_workbook(template_path)
        except FileNotFoundError:
            raise UserError("Excel template not found!")

        ws = wb.active

        ws["B4"] = f"Numele și prenumele: {self.employee_id.name or ''}"
        year_str = self.date.strftime("%Y") if self.date else ""
        ws["B6"] = f"Perioada de apreciere (anul): {year_str}"

        ws["C7"] = self.total
        ws["C8"] = dict(self._fields["final_score"].selection).get(self.final_score, "")

        # Fill answers in their respective columns
        for line in self.line_ids:
            answer = line.answer_id
            if answer and answer.column:
                cell = str(answer.column).upper()
                try:
                    ws[cell] = answer.score
                except Exception as e:
                    raise UserError(f"Invalid column '{cell}' in answer '{answer.name}': {str(e)}")

        # Save and export
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"Evaluation_{self.employee_id.name}_{datetime.date.today().isoformat()}.xlsx"
        self.write({
            "export_file": base64.b64encode(buf.read()),
            "export_filename": filename,
        })
        buf.close()

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{self._name}/{self.id}/export_file/{filename}?download=true",
            "target": "self",
        }


class HrEmployeeEvaluationLine(models.Model):
    _name = "hr.employee.evaluation.line"
    _description = "Evaluation Line"
    _order = "sequence, id"

    evaluation_id = fields.Many2one("hr.employee.evaluation", required=True, ondelete="cascade")
    question_id = fields.Many2one("hr.employee.evaluation.question", required=True, ondelete="restrict")
    sequence = fields.Integer(related="question_id.sequence", store=True)
    answer_id = fields.Many2one("hr.employee.evaluation.answer", string="Selected Answer")
    score = fields.Integer(compute="_compute_score", store=True)

    @api.depends("answer_id")
    def _compute_score(self):
        for line in self:
            line.score = line.answer_id.score if line.answer_id else 0


class HrEmployeeEvaluationQuestion(models.Model):
    _name = "hr.employee.evaluation.question"
    _description = "Evaluation Question"
    _order = "sequence, id"

    name = fields.Char(required=True)  # e.g. "Calitatea muncii"
    sequence = fields.Integer(default=10)
    active = fields.Boolean(default=True)
    answer_ids = fields.One2many("hr.employee.evaluation.answer", "question_id", string="Possible Answers")
    max_score = fields.Integer(default=5)


class HrEmployeeEvaluationAnswer(models.Model):
    _name = "hr.employee.evaluation.answer"
    _description = "Evaluation Answer"
    _order = "sequence, id"

    question_id = fields.Many2one("hr.employee.evaluation.question", required=True, ondelete="cascade")
    name = fields.Char(required=True)  # e.g. "Rezolvă sarcinile cu competență, cu atenție..."
    sequence = fields.Integer(default=10)
    score = fields.Integer(default=0)  # numeric value
    column = fields.Char(required=True)


class HrEmployeeEvaluationQuestionLead(models.Model):
    _name = "hr.employee.evaluation.question.lead"
    _description = "Evaluation Question"
    _order = "sequence, id"

    name = fields.Char(required=True)  # e.g. "Calitatea muncii"
    sequence = fields.Integer(default=10)
    active = fields.Boolean(default=True)
    answer_ids = fields.One2many("hr.employee.evaluation.answer.lead", "question_id", string="Possible Answers")
    max_score = fields.Integer(default=5)


class HrEmployeeEvaluationAnswerLead(models.Model):
    _name = "hr.employee.evaluation.answer.lead"
    _description = "Evaluation Answer"
    _order = "sequence, id"

    question_id = fields.Many2one("hr.employee.evaluation.question.lead", required=True, ondelete="cascade")
    name = fields.Char(required=True)  # e.g. "Rezolvă sarcinile cu competență, cu atenție..."
    sequence = fields.Integer(default=10)
    score = fields.Integer(default=0)  # numeric value
    column = fields.Char(required=True)





from odoo import models, fields, api

class HrEmployee(models.Model):
    _inherit = "hr.employee"

    evaluation_ids = fields.One2many(
        "hr.employee.evaluation", "employee_id", string="Evaluations"
    )
    evaluation_count = fields.Integer(compute="_compute_evaluation_count")

    def _compute_evaluation_count(self):
        mapped = self.env["hr.employee.evaluation"].read_group(
            [("employee_id", "in", self.ids)], ["employee_id"], ["employee_id"]
        )
        count_map = {m["employee_id"][0]: m["employee_id_count"] for m in mapped}
        for emp in self:
            emp.evaluation_count = count_map.get(emp.id, 0)
